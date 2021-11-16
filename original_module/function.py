import openpyxl as xl

def copy_cell(paste, copy, paste_row, paste_column, copy_row, copy_column):
                paste.cell(row = paste_row, column=paste_column).value=copy.cell(row=copy_row, column=copy_column).value

def copy(paste, copy, paste_top, paste_under, paste_left, paste_right, copy_row, copy_column):
    copy_row = copy_row - 1
    t = copy_column
    for i in range(paste_top, paste_under):
        copy_row = copy_row + 1
        copy_column = t
        for j in range(paste_left, paste_right):
            paste.cell(row=i, column=j).value = copy.cell(row = copy_row, column = copy_column).value
            copy_column += 1

def merge_copy(paste, copy, paste_top, paste_under, paste_left, copy_row, copy_column):
    k = 0
    for i in range(paste_top, paste_under):
        a = str(copy.cell(row = k + copy_row, column = copy_column).value)
        b = str(copy.cell(row = k + copy_row, column=copy_column + 1).value)
        c = a + b
        if(c != "NoneNone"):
            paste.cell(row=i, column=paste_left).value = c
        k += 1

def d_check(source,top,under,left,right):
    for i in range(top, under):
        for j in range(left, right):
            if(source.cell(row=i, column=j).value == "朝来"):
                source.cell(row=i, column=j).value = "×"
            elif(source.cell(row=i, column=j).value == "△"):
                source.cell(row=i, column=j).value = "×"
            elif(source.cell(row=i, column=j).value == "日帰"):
                source.cell(row=i, column=j).value = "×"
            elif(source.cell(row=i, column=j).value == "昼来"):
                source.cell(row=i, column=j).value = "×"
            elif(source.cell(row=i, column=j).value == "夜帰"):
                source.cell(row=i, column=j).value = "○"
            elif(source.cell(row=i, column=j).value == "夜来"):
                source.cell(row=i, column=j).value = None

def d_copy(paste, copy, paste_top, paste_under, paste_left, copy_row, copy_column):
    k = -1
    for i in range(paste_top, paste_under):
        k += 1
        for j in range(0, 6):
            paste.cell(row=i, column=2 * j + paste_left).value = copy.cell(row=copy_row + k, column=j + copy_column).value
            if(paste.cell(row=i, column=2 * j + paste_left).value == "朝来"):
                paste.cell(row=i, column=2 * j + paste_left).value = "×"
            elif(paste.cell(row=i, column=2 * j + paste_left).value == "△"):
                paste.cell(row=i, column=2 * j + paste_left).value = "×"
            elif(paste.cell(row=i, column=2 * j + paste_left).value == "日帰"):
                paste.cell(row=i, column=2 * j + paste_left).value = "×"
            elif(paste.cell(row=i, column=2 * j + paste_left).value == "昼来"):
                paste.cell(row=i, column=2 * j + paste_left).value = "×"
            elif(paste.cell(row=i, column=2 * j + paste_left).value == "夜帰"):
                paste.cell(row=i, column=2 * j + paste_left).value = "○"
            elif(paste.cell(row=i, column=2 * j + paste_left).value == "夜来"):
                paste.cell(row=i, column=2 * j + paste_left).value = None

def k_check_copy(paste, copy, top, under, k_column):
    a = 7
    for i in range(top, under):
        if(a==29):
            break
        paste.cell(row=i, column=k_column).value=copy.cell(row=a, column=5).value
        if(paste.cell(row=i, column=k_column).value == "指導員"):
            paste.cell(row=i, column=k_column).value="指導員"
        elif(paste.cell(row=i, column=k_column).value == "助教"or paste.cell(row=i, column=k_column).value =="助教官"):
            paste.cell(row=i, column=k_column).value="助教官"
        elif(paste.cell(row=i, column=k_column).value == "見習い"or paste.cell(row=i, column=k_column).value =="見習教官"):
            paste.cell(row=i, column=k_column).value="見習教官"    
        a += 1

def syarei_check(source, top, under, left, right):
    for i in range(top, under):
        for j in range(left, right):
            if(source.cell(row=i, column=j).value == "○" or source.cell(row=i, column=j).value == "×"):
                source.cell(row=i, column=j).value = 1
