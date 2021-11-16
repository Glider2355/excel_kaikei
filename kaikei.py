# coding: utf-8
import openpyxl as xl
from original_module.function import *
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox

# 以下内部処理

#GUI
def GUI():
    def OpenFileDlg(tbox):
        ftype = [('','*')]
        dir = '.'
        filename = filedialog.askopenfilename(filetypes = ftype, initialdir = dir)
        tbox.insert(0, filename)

    root = tk.Tk()
    root.title('合宿会計')
    root.geometry("400x200")

    label = tk.Label(root, text = 'ピストファイル')
    label.place(x = 30, y = 10)
    pisuto = tk.Entry(root, width = 40)
    pisuto.place(x = 30, y = 30)
    fdlg_button = tk.Button(root, text = 'ファイル選択', command = lambda: OpenFileDlg(pisuto))
    fdlg_button.place(x = 280, y = 30)

    label_mode = tk.Label(root, text = 'モード選択')
    label_mode.place(x = 30, y = 50)
    OptionList = ["木曽川", "大野", "福井"]
    variable = tk.StringVar(root)
    variable.set(OptionList[0])

    opt = tk.OptionMenu(root, variable, *OptionList)
    opt.config(width = 30, font = ('Helvetica', 12))
    opt.place(x = 30, y = 70)

    label_xlsx = tk.Label(root, text = '.xlsx')
    label_xlsx.place(x = 280, y = 140)
    label_name = tk.Label(root, text = 'ファイル名')
    label_name.place(x = 30, y = 110)
    name = tk.Entry(root, width = 40)
    name.place(x = 30, y = 140)

    make_button = tk.Button(root, text = '作成', command =lambda: main(variable.get(),name.get() + ".xlsx",pisuto.get()))
    make_button.place(x = 170, y = 170)

    root.mainloop()


def main(place,name,pisuto):
    # 木曽川
    if(place == "木曽川"):

        # 合宿フォーマットファイル名
        temp = 'format/木曽川会計 改定.xlsx'
        # wb1=合宿会計シート
        wb1 = xl.load_workbook(filename=temp)
        # wb2=ピスト出入り
        wb2 = xl.load_workbook(filename=pisuto)
        deiri1 = wb1.worksheets[2]
        deiri2 = wb2.worksheets[2]
        deiri1_2 = wb1.worksheets[7]
        kdeiri = wb2.worksheets[1]
        syarei = wb1.worksheets[4]
        kozin = wb1.worksheets[0]

        wb3 = xl.load_workbook(filename='format/個人データ.xlsx')
        data = wb3.worksheets[0]

        #個人データコピー
        for row in data:
            for cell in row:
                kozin[cell.coordinate].value = cell.value
    
        # 出入り入力(訓練報告書2)
        for i in range(3, 33):
            deiri1.cell(row=i, column=2).value = deiri2.cell(
                row=i + 4, column=2).value
            deiri1.cell(row=i, column=3).value = deiri2.cell(
                row=i + 4, column=3).value
            for j in range(0, 7):
                deiri1.cell(row=i, column=2 * j +
                            5).value = deiri2.cell(row=i+4, column=j+9).value
                if(deiri1.cell(row=i, column=2*j + 5).value == "朝来"):
                    deiri1.cell(row=i, column=2*j + 5).value = "×"
                elif(deiri1.cell(row=i, column=2*j + 5).value == "△"):
                    deiri1.cell(row=i, column=2*j + 5).value = "×"
                elif(deiri1.cell(row=i, column=2*j + 5).value == "日帰"):
                    deiri1.cell(row=i, column=2*j + 5).value = "×"
                if(deiri1.cell(row=i, column=2*j + 5).value == "昼来"):
                    deiri1.cell(row=i, column=2*j + 5).value = "×"
                elif(deiri1.cell(row=i, column=2*j + 5).value == "夜帰"):
                    deiri1.cell(row=i, column=2*j + 5).value = "○"
                elif(deiri1.cell(row=i, column=2*j + 5).value == "夜来"):
                    deiri1.cell(row=i, column=2*j + 5).value = None

        # 出入り入力(訓練報告書2(1)参加訓練生)
        for i in range(3, 29):
            deiri1_2.cell(row=i, column=2).value = deiri2.cell(
                row=i+34, column=2).value
            deiri1_2.cell(row=i, column=3).value = deiri2.cell(
                row=i+34, column=3).value
            for j in range(0, 7):
                deiri1_2.cell(row=i, column=2*j+5).value = deiri2.cell(
                    row=i+34, column=j+9).value
                if(deiri1_2.cell(row=i, column=2*j + 5).value == "朝来"):
                    deiri1_2.cell(row=i, column=2*j + 5).value = "×"
                elif(deiri1_2.cell(row=i, column=2*j + 5).value == "△"):
                    deiri1_2.cell(row=i, column=2*j + 5).value = "×"
                elif(deiri1_2.cell(row=i, column=2*j + 5).value == "日帰"):
                    deiri1_2.cell(row=i, column=2*j + 5).value = "×"
                elif(deiri1_2.cell(row=i, column=2*j + 5).value == "昼来"):
                    deiri1_2.cell(row=i, column=2*j + 5).value = "×"
                elif(deiri1_2.cell(row=i, column=2*j + 5).value == "夜帰"):
                    deiri1_2.cell(row=i, column=2*j + 5).value = "○"
                elif(deiri1_2.cell(row=i, column=2*j + 5).value == "夜来"):
                    deiri1_2.cell(row=i, column=2*j + 5).value = None

        # 名前入力(訓練報告書2(1)参加教官)
        i = 7
        for k in range(3, 29):
            if(deiri1_2.cell(row=k, column=2).value == None):
                deiri1_2.cell(row=k, column=2).value = kdeiri.cell(
                    row=i, column=2).value
                deiri1_2.cell(row=k, column=3).value = kdeiri.cell(
                    row=i, column=3).value
                for j in range(0, 7):
                    deiri1_2.cell(row=k, column=2*j+5).value = kdeiri.cell(
                        row=i, column=j+8).value
                    if(deiri1_2.cell(row=k, column=2*j + 5).value == "朝来"):
                        deiri1_2.cell(row=k, column=2*j + 5).value = "×"
                    elif(deiri1_2.cell(row=k, column=2*j + 5).value == "△"):
                        deiri1_2.cell(row=k, column=2*j + 5).value = "×"
                    elif(deiri1_2.cell(row=k, column=2*j + 5).value == "日帰"):
                        deiri1_2.cell(row=k, column=2*j + 5).value = "×"
                    elif(deiri1_2.cell(row=k, column=2*j + 5).value == "昼来"):
                        deiri1_2.cell(row=k, column=2*j + 5).value = "×"
                    elif(deiri1_2.cell(row=k, column=2*j + 5).value == "夜帰"):
                        deiri1_2.cell(row=k, column=2*j + 5).value = "○"
                    elif(deiri1_2.cell(row=k, column=2*j + 5).value == "夜来"):
                        deiri1_2.cell(row=k, column=2*j + 5).value = None
                i = i+1

        #謝礼コピー
        merge_copy(syarei,kdeiri,4,18,1,7,2)
        k_check_copy(syarei,kdeiri,4,18,2)
        copy(syarei,kdeiri,4,16,3,9,7,8)
        d_check(syarei,4,16,3,9)
        syarei_check(syarei,4,16,3,9)

        wb1.save(name)

    # 大野
    elif(place == "大野"):
        # 合宿フォーマットファイル名
        temp = 'format/大野会計 改定.xlsx'
        # wb1=合宿会計シート
        wb1 = xl.load_workbook(filename=temp)
        # wb2=ピスト出入り
        wb2 = xl.load_workbook(filename=pisuto)
        deiri1 = wb1.worksheets[2]
        deiri2 = wb2.worksheets[2]
        deiri1_2 = wb1.worksheets[6]
        kdeiri = wb2.worksheets[1]
        deiri1_3 = wb1.worksheets[8]
        kozin = wb1.worksheets[0]

        wb3 = xl.load_workbook(filename='format/個人データ.xlsx')
        data = wb3.worksheets[0]

        #個人データコピー
        for row in data:
            for cell in row:
                kozin[cell.coordinate].value = cell.value

        # 出入り入力(訓練報告書2)
        for i in range(3, 33):
            deiri1.cell(row=i, column=2).value = deiri2.cell(
                row=i+4, column=2).value
            deiri1.cell(row=i, column=3).value = deiri2.cell(
                row=i+4, column=3).value
            for j in range(0, 7):
                deiri1.cell(row=i, column=2*j +
                            5).value = deiri2.cell(row=i+4, column=j+9).value
                if(deiri1.cell(row=i, column=2*j + 5).value == "朝来"):
                    deiri1.cell(row=i, column=2*j + 5).value = "×"
                elif(deiri1.cell(row=i, column=2*j + 5).value == "△"):
                    deiri1.cell(row=i, column=2*j + 5).value = "×"
                elif(deiri1.cell(row=i, column=2*j + 5).value == "日帰"):
                    deiri1.cell(row=i, column=2*j + 5).value = "×"
                if(deiri1.cell(row=i, column=2*j + 5).value == "昼来"):
                    deiri1.cell(row=i, column=2*j + 5).value = "×"
                elif(deiri1.cell(row=i, column=2*j + 5).value == "夜帰"):
                    deiri1.cell(row=i, column=2*j + 5).value = "○"
                elif(deiri1.cell(row=i, column=2*j + 5).value == "夜来"):
                    deiri1.cell(row=i, column=2*j + 5).value = None

        # 出入り入力(訓練報告書2(1))
        i = 7
        for k in range(4, 17):
            deiri1_2.cell(row=k, column=1).value = kdeiri.cell(
                row=i, column=2).value
            deiri1_2.cell(row=k, column=2).value = kdeiri.cell(
                row=i, column=3).value
            deiri1_2.cell(row=k, column=3).value = kdeiri.cell(
                row=i, column=5).value
            if(deiri1_2.cell(row=k, column=3).value == "指導員"):
                deiri1_2.cell(row=k, column=3).value="指導員"
            elif(deiri1_2.cell(row=k, column=3).value == "助教"or deiri1_2.cell(row=k, column=3).value =="助教官"):
                deiri1_2.cell(row=k, column=3).value="助教官"
            elif(deiri1_2.cell(row=k, column=3).value == "見習い"or deiri1_2.cell(row=k, column=3).value =="見習教官"):
                deiri1_2.cell(row=k, column=3).value="見習教官"
            for j in range(0, 7):
                deiri1_2.cell(row=k, column=j+4).value = kdeiri.cell(
                    row=i, column=j+8).value
                if(deiri1_2.cell(row=k, column=j + 4).value == "朝来"):
                    deiri1_2.cell(row=k, column=j + 4).value = "×"
                elif(deiri1_2.cell(row=k, column=2*j + 4).value == "△"):
                        deiri1_2.cell(row=k, column=2*j + 4).value = "×"
                elif(deiri1_2.cell(row=k, column=2*j + 4).value == "日帰"):
                    deiri1_2.cell(row=k, column=2*j + 4).value = "×"
                elif(deiri1_2.cell(row=k, column=j + 4).value == "昼来"):
                    deiri1_2.cell(row=k, column=j + 4).value = "×"
                elif(deiri1_2.cell(row=k, column=j + 4).value == "夜帰"):
                    deiri1_2.cell(row=k, column=j + 4).value = "○"
                elif(deiri1_2.cell(row=k, column=j + 4).value == "夜来"):
                    deiri1_2.cell(row=k, column=j + 4).value = None
                if(deiri1_2.cell(row=k, column=j + 4).value == "○" or deiri1_2.cell(row=k, column=j + 4).value =="×"):
                    deiri1_2.cell(row=k, column=j + 4).value=1
            i = i+1

        #出入りコピー
        copy(deiri1_3,deiri2,3,33,2,4,7,2)
        copy(deiri1_3,deiri2,3,33,4,10,7,8)
        for k in range(3,39):
            if(deiri1_3.cell(row=k, column=2).value == None):
                j=k
                break
        copy(deiri1_3,kdeiri,j,39,2,4,7,2)
        copy(deiri1_3,kdeiri,j,39,4,11,7,7)
        wb1.save(name)

    # 福井
    if(place == "福井"):
        # 合宿フォーマットファイル名
        temp = 'format/福井白紙会計.xlsx'
        # wb1=合宿会計シート
        wb1 = xl.load_workbook(filename=temp)
        # wb2=ピスト出入り
        wb2 = xl.load_workbook(filename=pisuto)
        deiri1 = wb1.worksheets[6]
        deiri2 = wb2.worksheets[2]
        kyoukan = wb2.worksheets[1]
        syarei = wb1.worksheets[4]
        kozin = wb1.worksheets[0]

        wb3 = xl.load_workbook(filename='format/個人データ.xlsx')
        data = wb3.worksheets[0]

        #個人データコピー
        for row in data:
            for cell in row:
                kozin[cell.coordinate].value = cell.value

        copy(deiri1,deiri2,3,42,2,4,7,2)
        d_copy(deiri1,deiri2,3,42,5,7,9)
        merge_copy(syarei,kyoukan,4,18,1,7,2)
        k_check_copy(syarei,kyoukan,4,18,2)
        copy(syarei,kyoukan,4,16,3,9,7,8)
        syarei_check(syarei,4,16,3,9)

        wb1.save(name)
GUI()
